#!/usr/bin/env python3
"""
Endeavor Science (AETH-1) - 5-Year Financial Projections Model
===============================================================
Comprehensive SaaS financial model with 8 sheets:
1. Executive Summary
2. User Forecast (60-month)
3. Revenue Model (MRR/ARR)
4. P&L Statement (5-Year)
5. Headcount Plan
6. Cash Flow Statement
7. Key Metrics & KPIs
8. Charts

Author: Z.ai Financial Modeling
"""

import sys
import os

# Setup paths for xlsx skill imports
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
from openpyxl.comments import Comment

# Import design system from base.py
from templates.base import (
    FONT_NAME, HEADER_BOLD, PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    HEADER_TEXT, CHART_COLORS,
    CF_POSITIVE_FILL, CF_POSITIVE_FONT,
    CF_NEGATIVE_FILL, CF_NEGATIVE_FONT,
    CF_WARNING_FILL, CF_WARNING_FONT,
    font_title, font_header, font_subheader, font_body, font_caption, font_kpi, font_kpi_label,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    create_bar_chart, create_line_chart, create_pie_chart,
    setup_chart_titles, apply_chart_colors, apply_pie_colors,
    auto_fit_columns, COLUMN_WIDTHS, ROW_HEIGHTS, FORMATS
)

# ============================================================
# FINANCE-SPECIFIC FORMATS
# ============================================================
FINANCE_FORMATS = {
    'currency': '$#,##0;($#,##0);"-"',
    'currency_k': '$#,##0,"K";($#,##0,"K");"-"',
    'currency_mm': '$#,##0.0,,"M";($#,##0.0,,"M");"-"',
    'pct': '0.0%;(0.0%);"-"',
    'pct_decimal': '0.00%;(0.00%);"-"',
    'multiple': '0.0"x";(0.0"x");"-"',
    'integer': '#,##0;(#,##0);"-"',
    'decimal': '#,##0.00;(#,##0.00);"-"',
}

# ============================================================
# COMPANY ASSUMPTIONS
# ============================================================
COMPANY = {
    'name': 'Endeavor Science (AETH-1)',
    'tagline': 'The GitHub for Scientific Computing',
    'model': 'Freemium SaaS Platform'
}

# User Growth Targets (Year-End)
USER_TARGETS = {
    1: {'total': 100000, 'paid': 5000},
    2: {'total': 350000, 'paid': 15000},
    3: {'total': 1000000, 'paid': 50000},
    4: {'total': 4000000, 'paid': 150000},
    5: {'total': 10000000, 'paid': 500000},
}

# Pricing
PRICING = {
    'professional_monthly': 50,
    'professional_annual': 600,
    'enterprise_monthly': 500,
    'enterprise_annual': 6000,
    'professional_mix': 0.70,  # 70% of paid users
    'enterprise_mix': 0.30,   # 30% of paid users
}

# Churn Rates
CHURN = {
    'enterprise': 0.08,   # 8% annual churn
    'professional': 0.15, # 15% annual churn
}

# Operating Expense as % of Revenue
OPEX_PCT = {
    'sales_marketing_start': 0.40,
    'sales_marketing_end': 0.25,
    'rd_start': 0.50,
    'rd_end': 0.30,
    'ga_pct': 0.15,
}

# Headcount Assumptions
HEADCOUNT = {
    'engineering_avg_salary': 180000,
    'sales_avg_salary': 120000,
    'marketing_avg_salary': 100000,
    'ops_avg_salary': 90000,
    'ga_avg_salary': 110000,
    'benefits_burden': 0.20,  # 20% benefits burden
}

# Target ARR validation
TARGET_ARR = {1: 5000000, 3: 50000000, 5: 500000000}


def create_workbook():
    """Create and return workbook with all sheets."""
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Create all sheets
    sheet_names = [
        'Executive Summary',
        'User Forecast', 
        'Revenue Model',
        'P&L Statement',
        'Headcount Plan',
        'Cash Flow',
        'Key Metrics & KPIs',
        'Charts'
    ]
    
    for name in sheet_names:
        wb.create_sheet(title=name)
    
    # Remove default sheet
    wb.remove(default_sheet)
    
    return wb


def apply_finance_style(cell, is_input=False, is_reference=False):
    """Apply finance-specific styling to cell."""
    cell.font = Font(name=FONT_NAME, size=11, color=NEUTRAL_900)
    if is_input:
        cell.font = Font(name=FONT_NAME, size=11, color='0000FF')  # Blue for inputs
    elif is_reference:
        cell.font = Font(name=FONT_NAME, size=11, color='008000')  # Green for references


def write_executive_summary(wb):
    """Sheet 1: Executive Summary with KPI dashboard."""
    ws = wb['Executive Summary']
    last_col = 7  # A-G columns used
    
    setup_sheet(ws, title=f"{COMPANY['name']} - Executive Summary", last_col=last_col)
    
    # Company Info Section
    row = 4
    ws.cell(row=row, column=2, value="Company Overview")
    ws.cell(row=row, column=2).font = font_subheader()
    row += 1
    ws.cell(row=row, column=2, value=f"Tagline: {COMPANY['tagline']}")
    ws.cell(row=row, column=2).font = font_body()
    row += 1
    ws.cell(row=row, column=2, value=f"Business Model: {COMPANY['model']}")
    ws.cell(row=row, column=2).font = font_body()
    
    # 5-Year Summary Table Header
    row += 2
    headers = ['Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col_idx, value=header)
    style_header_row(ws, row, 2, len(headers) + 1)
    
    # Data rows with formulas referencing other sheets
    metrics = [
        ('Total Users', "='User Forecast'!C67", "='User Forecast'!G67", "='User Forecast'!K67", "='User Forecast'!O67", "='User Forecast'!S67"),
        ('Paid Users', "='User Forecast'!D67", "='User Forecast'!H67", "='User Forecast'!L67", "='User Forecast'!P67", "='User Forecast'!T67"),
        ('Total Revenue ($M)', "='P&L Statement'!G12/1000000", "='P&L Statement'!H12/1000000", "='P&L Statement'!I12/1000000", "='P&L Statement'!J12/1000000", "='P&L Statement'!K12/1000000"),
        ('ARR ($M)', "='Revenue Model'!B47/1000000", "='Revenue Model'!C47/1000000", "='Revenue Model'!D47/1000000", "='Revenue Model'!E47/1000000", "='Revenue Model'!F47/1000000"),
        ('Gross Margin %', "='P&L Statement'!G17", "='P&L Statement'!H17", "='P&L Statement'!I17", "='P&L Statement'!J17", "='P&L Statement'!K17"),
        ('EBITDA ($M)', "='P&L Statement'!G26/1000000", "='P&L Statement'!H26/1000000", "='P&L Statement'!I26/1000000", "='P&L Statement'!J26/1000000", "='P&L Statement'!K26/1000000"),
        ('Net Income ($M)', "='P&L Statement'!G32/1000000", "='P&L Statement'!H32/1000000", "='P&L Statement'!I32/1000000", "='P&L Statement'!J32/1000000", "='P&L Statement'!K32/1000000"),
    ]
    
    for i, metric_data in enumerate(metrics):
        row_num = row + 1 + i
        metric_name = metric_data[0]
        ws.cell(row=row_num, column=2, value=metric_name)
        ws.cell(row=row_num, column=2).font = font_body()
        
        for col_idx, formula in enumerate(metric_data[1:], start=3):
            cell = ws.cell(row=row_num, column=col_idx, value=formula)
            apply_finance_style(cell)
            
            # Apply number formats based on metric type
            if 'Users' in metric_name:
                cell.number_format = '#,##0'
            elif '%' in metric_name:
                cell.number_format = FINANCE_FORMATS['pct']
            else:
                cell.number_format = FINANCE_FORMATS['currency_mm']
        
        style_data_row(ws, row_num, 2, last_col, i)
    
    # Style totals row (last metric row)
    total_row = row + len(metrics)
    style_total_row(ws, total_row, 2, last_col)
    
    # Growth Rates Section
    growth_row = total_row + 2
    ws.cell(row=growth_row, column=2, value="YoY Growth Rates")
    ws.merge_cells(start_row=growth_row, start_column=2, end_row=growth_row, end_column=last_col)
    ws.cell(row=growth_row, column=2).font = font_subheader()
    
    growth_headers = ['Growth Metric', 'Y1→Y2', 'Y2→Y3', 'Y3→Y4', 'Y4→Y5']
    growth_row += 1
    for col_idx, header in enumerate(growth_headers, start=2):
        ws.cell(row=growth_row, column=col_idx, value=header)
    style_header_row(ws, growth_row, 2, 6)
    
    growth_metrics = [
        ('User Growth %', '=IFERROR((D7-C7)/C7,0)', '=IFERROR((E7-D7)/D7,0)', '=IFERROR((F7-E7)/E7,0)', '=IFERROR((G7-F7)/F7,0)'),
        ('Revenue Growth %', '=IFERROR((D9-C9)/C9,0)', '=IFERROR((E9-D9)/D9,0)', '=IFERROR((F9-E9)/E9,0)', '=IFERROR((G9-F9)/F9,0)'),
        ('ARR Growth %', '=IFERROR((D10-C10)/C10,0)', '=IFERROR((E10-D10)/D10,0)', '=IFERROR((F10-E10)/E10,0)', '=IFERROR((G10-F10)/F10,0)'),
    ]
    
    for i, gm in enumerate(growth_metrics):
        gr_row = growth_row + 1 + i
        ws.cell(row=gr_row, column=2, value=gm[0])
        for col_idx, formula in enumerate(gm[1:], start=3):
            cell = ws.cell(row=gr_row, column=col_idx, value=formula)
            cell.number_format = FINANCE_FORMATS['pct']
            apply_finance_style(cell)
        style_data_row(ws, gr_row, 2, 6, i)
    
    # KPI Dashboard Section
    kpi_row = growth_row + len(growth_metrics) + 3
    ws.cell(row=kpi_row, column=2, value="Key Performance Indicators")
    ws.merge_cells(start_row=kpi_row, start_column=2, end_row=kpi_row, end_column=last_col)
    ws.cell(row=kpi_row, column=2).font = font_subheader()
    
    kpis = [
        ('Target ARR Y1', f"${TARGET_ARR[1]:,.0f}"),
        ('Target ARR Y3', f"${TARGET_ARR[3]:,.0f}"),
        ('Target ARR Y5', f"${TARGET_ARR[5]:,.0f}"),
        ('Professional Price/Mo', f"${PRICING['professional_monthly']:,}/mo"),
        ('Enterprise Price/Mo', f"${PRICING['enterprise_monthly']:,}/mo"),
        ('Enterprise Churn', f"{CHURN['enterprise']*100:.0f}%"),
        ('Professional Churn', f"{CHURN['professional']*100:.0f}%"),
    ]
    
    kpi_row += 1
    for i, (kpi_name, kpi_val) in enumerate(kpis):
        ws.cell(row=kpi_row + i, column=2, value=kpi_name)
        ws.cell(row=kpi_row + i, column=3, value=kpi_val)
        ws.cell(row=kpi_row + i, column=2).font = font_body()
        ws.cell(row=kpi_row + i, column=3).font = font_body()
    
    # Set column widths
    ws.column_dimensions['B'].width = 22
    for col in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    return ws


def write_user_forecast(wb):
    """Sheet 2: User Forecast with 60-month projections."""
    ws = wb['User Forecast']
    last_col = 22  # A-V columns for monthly data
    
    setup_sheet(ws, title="User Forecast - 60 Month Projection", last_col=last_col)
    
    # Assumptions Section
    row = 4
    ws.cell(row=row, column=2, value="Growth Assumptions")
    ws.cell(row=row, column=2).font = font_subheader()
    
    assumptions = [
        ('Year 1 End Users', USER_TARGETS[1]['total']),
        ('Year 2 End Users', USER_TARGETS[2]['total']),
        ('Year 3 End Users', USER_TARGETS[3]['total']),
        ('Year 4 End Users', USER_TARGETS[4]['total']),
        ('Year 5 End Users', USER_TARGETS[5]['total']),
        ('Paid Conversion Rate Y1', USER_TARGETS[1]['paid']/USER_TARGETS[1]['total']),
        ('Paid Conversion Rate Y5', USER_TARGETS[5]['paid']/USER_TARGETS[5]['total']),
        ('Enterprise Churn (Annual)', CHURN['enterprise']),
        ('Professional Churn (Annual)', CHURN['professional']),
    ]
    
    row += 1
    for i, (name, val) in enumerate(assumptions):
        ws.cell(row=row+i, column=2, value=name)
        cell = ws.cell(row=row+i, column=3, value=val)
        apply_finance_style(cell, is_input=True)
        if 'Rate' in name or 'Churn' in name:
            cell.number_format = FINANCE_FORMATS['pct']
        else:
            cell.number_format = FINANCE_FORMATS['integer']
    
    # Monthly Headers (60 months)
    header_row = row + len(assumptions) + 2
    months = []
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    for year in range(1, 6):
        for month in month_names:
            months.append(f"Y{year}-{month}")
    
    # Write headers - grouped by year with sub-columns
    headers = ['Metric']
    year_cols = {}  # Track column positions per year
    
    col = 3
    for year in range(1, 6):
        year_start_col = col
        for m in range(12):
            headers.append(f"Y{year}-M{m+1}")
            col += 1
        year_cols[year] = (year_start_col, col - 1)  # Start and end column for each year
    
    for col_idx, h in enumerate(headers, start=2):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_header_row(ws, header_row, 2, col)
    
    # Data rows
    data_rows = [
        'New Free Users',
        'Churned Free Users',
        'Total Free Users',
        'New Paid Users (Pro)',
        'New Paid Users (Ent)',
        'Churned Paid Users (Pro)',
        'Churned Paid Users (Ent)',
        'Total Paid Users (Pro)',
        'Total Paid Users (Ent)',
        'Total Paid Users',
        'Total Users',
    ]
    
    # Starting values
    start_free = 0
    start_paid_pro = 0
    start_paid_ent = 0
    
    monthly_churn_pro = 1 - (1 - CHURN['professional']) ** (1/12)
    monthly_churn_ent = 1 - (1 - CHURN['enterprise']) ** (1/12)
    
    # Calculate monthly targets using interpolation/growth curves
    def get_monthly_target(year, month, target_yearend, prev_end=0):
        """Calculate monthly user count with smooth growth curve."""
        # Use exponential growth within each year
        year_progress = month / 12
        prev_year_target = USER_TARGETS.get(year-1, {}).get('total', 0) if year > 1 else 0
        
        # Smooth S-curve-like growth
        if prev_year_target > 0:
            base = prev_year_target
        else:
            base = 0
            
        target = target_yearend
        # Exponential interpolation
        if base > 0:
            growth_factor = (target / base) ** (1/12)
            return int(base * (growth_factor ** month))
        else:
            # First year - linear ramp up
            return int(target * (month / 12) * 1.5)  # Front-load slightly
    
    for r_idx, row_name in enumerate(data_rows):
        row_num = header_row + 1 + r_idx
        ws.cell(row=row_num, column=2, value=row_name)
        ws.cell(row=row_num, column=2).font = font_body()
        
        for month_idx in range(60):
            year = (month_idx // 12) + 1
            month_in_year = (month_idx % 12) + 1
            col_idx = 3 + month_idx
            cell = ws.cell(row=row_num, column=col_idx)
            
            # Previous month column letter
            if month_idx == 0:
                prev_col = None
                prev_free = 0
                prev_pro = 0
                prev_ent = 0
            else:
                prev_col = get_column_letter(col_idx - 1)
                prev_free = f"{prev_col}{header_row + 3}"  # Total Free Users row
                prev_pro = f"{prev_col}{header_row + 9}"   # Total Pro row
                prev_ent = f"{prev_col}{header_row + 10}"  # Total Ent row
            
            current_col = get_column_letter(col_idx)
            
            if row_name == 'New Free Users':
                # New free users - calculated to hit year-end targets
                if month_in_year == 1:
                    if year == 1:
                        # First month of first year
                        val = int(USER_TARGETS[year]['total'] / 12 * 0.3)
                    else:
                        prev_year_total = USER_TARGETS[year-1]['total']
                        val = int((USER_TARGETS[year]['total'] - prev_year_total) / 12 * 0.4)
                else:
                    # Growth pattern
                    base_new = USER_TARGETS[year]['total'] // 12
                    seasonal_factor = 1.0 + 0.1 * (1 if month_in_year in [9, 10] else (-0.1 if month_in_year in [6, 7] else 0))
                    val = int(base_new * seasonal_factor)
                cell.value = val
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'Churned Free Users':
                # Assume 5% annual churn for free users (~0.4% monthly)
                if month_idx > 0:
                    cell.value = f"=ROUND({prev_free}*0.004,0)"
                else:
                    cell.value = 0
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'Total Free Users':
                if month_idx == 0:
                    cell.value = f"={current_col}{header_row+1}"  # New Free
                else:
                    cell.value = f"={prev_free}+{current_col}{header_row+1}-{current_col}{header_row+2}"
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'New Paid Users (Pro)':
                # 70% of new paid are Pro
                target_paid = USER_TARGETS[year]['paid']
                pro_target = int(target_paid * PRICING['professional_mix'])
                if month_in_year == 1:
                    val = max(1, int(pro_target / 12 * 0.3)) if year == 1 else int(pro_target / 12 * 0.35)
                else:
                    val = max(1, int(pro_target / 12))
                cell.value = val
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'New Paid Users (Ent)':
                # 30% of new paid are Enterprise
                target_paid = USER_TARGETS[year]['paid']
                ent_target = int(target_paid * PRICING['enterprise_mix'])
                if month_in_year == 1:
                    val = max(1, int(ent_target / 12 * 0.25)) if year == 1 else int(ent_target / 12 * 0.3)
                else:
                    val = max(1, int(ent_target / 12))
                cell.value = val
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'Churned Paid Users (Pro)':
                if month_idx > 0:
                    cell.value = f"=ROUND({prev_pro}*{monthly_churn_pro:.4f},0)"
                else:
                    cell.value = 0
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'Churned Paid Users (Ent)':
                if month_idx > 0:
                    cell.value = f"=ROUND({prev_ent}*{monthly_churn_ent:.4f},0)"
                else:
                    cell.value = 0
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'Total Paid Users (Pro)':
                if month_idx == 0:
                    cell.value = f"={current_col}{header_row+5}"
                else:
                    cell.value = f"={prev_pro}+{current_col}{header_row+5}-{current_col}{header_row+7}"
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'Total Paid Users (Ent)':
                if month_idx == 0:
                    cell.value = f"={current_col}{header_row+6}"
                else:
                    cell.value = f"={prev_ent}+{current_col}{header_row+6}-{current_col}{header_row+8}"
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'Total Paid Users':
                cell.value = f"={current_col}{header_row+9}+{current_col}{header_row+10}"
                cell.number_format = FINANCE_FORMATS['integer']
                
            elif row_name == 'Total Users':
                cell.value = f"={current_col}{header_row+3}+{current_col}{header_row+11}"
                cell.number_format = FINANCE_FORMATS['integer']
            
            apply_finance_style(cell)
        
        style_data_row(ws, row_num, 2, col, r_idx)
    
    # Year-end summary section
    summary_row = header_row + len(data_rows) + 2
    ws.cell(row=summary_row, column=2, value="Year-End Summary")
    ws.cell(row=summary_row, column=2).font = font_subheader()
    
    summary_headers = ['Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    summary_row += 1
    for col_idx, h in enumerate(summary_headers, start=2):
        ws.cell(row=summary_row, column=col_idx, value=h)
    style_header_row(ws, summary_row, 2, 7)
    
    # Year-end totals reference December of each year
    dec_cols = [get_column_letter(3 + 11), get_column_letter(3 + 23), 
                get_column_letter(3 + 35), get_column_letter(3 + 47), get_column_letter(3 + 59)]
    
    summary_metrics = [
        ('Total Users', header_row + 12),
        ('Total Paid Users', header_row + 11),
        ('Pro Users', header_row + 9),
        ('Enterprise Users', header_row + 10),
    ]
    
    for i, (metric, src_row) in enumerate(summary_metrics):
        s_row = summary_row + 1 + i
        ws.cell(row=s_row, column=2, value=metric)
        for j, dec_col in enumerate(dec_cols):
            cell = ws.cell(row=s_row, column=3+j, value=f"={dec_col}{src_row}")
            cell.number_format = FINANCE_FORMATS['integer']
            apply_finance_style(cell)
        style_data_row(ws, s_row, 2, 7, i)
    
    style_total_row(ws, summary_row + len(summary_metrics), 2, 7)
    
    # Set column widths
    ws.column_dimensions['B'].width = 24
    for c in range(3, col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9
    
    return ws


def write_revenue_model(wb):
    """Sheet 3: Revenue Model with MRR/ARR breakdown."""
    ws = wb['Revenue Model']
    last_col = 8
    
    setup_sheet(ws, title="Revenue Model - MRR & ARR Analysis", last_col=last_col)
    
    # Pricing Assumptions
    row = 4
    ws.cell(row=row, column=2, value="Pricing & Tier Assumptions")
    ws.cell(row=row, column=2).font = font_subheader()
    
    pricing_assumptions = [
        ('Professional Monthly Price', PRICING['professional_monthly'], FINANCE_FORMATS['currency']),
        ('Enterprise Monthly Price', PRICING['enterprise_monthly'], FINANCE_FORMATS['currency']),
        ('Professional Annual Value', PRICING['professional_annual'], FINANCE_FORMATS['currency']),
        ('Enterprise Annual Value', PRICING['enterprise_annual'], FINANCE_FORMATS['currency']),
        ('Professional Mix (%)', PRICING['professional_mix'], FINANCE_FORMATS['pct']),
        ('Enterprise Mix (%)', PRICING['enterprise_mix'], FINANCE_FORMATS['pct']),
    ]
    
    row += 1
    for i, (name, val, fmt) in enumerate(pricing_assumptions):
        ws.cell(row=row+i, column=2, value=name)
        cell = ws.cell(row=row+i, column=3, value=val)
        apply_finance_style(cell, is_input=True)
        cell.number_format = fmt
    
    # Main Revenue Table
    table_row = row + len(pricing_assumptions) + 2
    headers = ['Revenue Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    for col_idx, h in enumerate(headers, start=2):
        ws.cell(row=table_row, column=col_idx, value=h)
    style_header_row(ws, table_row, 2, 7)
    
    # Revenue calculations
    revenue_items = [
        # (Row Name, Formula pattern, Number Format, Is Total Row)
        ('Professional Users (Avg)', 
         ["='User Forecast'!D67", "='User Forecast'!H67", "='User Forecast'!L67", "='User Forecast'!P67", "='User Forecast'!T67"],
         FINANCE_FORMATS['integer'], False),
        ('Enterprise Users (Avg)', 
         ["='User Forecast'!E68", "='User Forecast'!I68", "='User Forecast'!M68", "='User Forecast'!Q68", "='User Forecast'!U68"],
         FINANCE_FORMATS['integer'], False),
        ('Professional MRR', 
         ["=B{r}*{pro_price}".format(r=table_row+1, pro_price=PRICING['professional_monthly'])] * 5,
         FINANCE_FORMATS['currency'], False),
        ('Enterprise MRR', 
         ["=B{r}*{ent_price}".format(r=table_row+2, ent_price=PRICING['enterprise_monthly'])] * 5,
         FINANCE_FORMATS['currency'], False),
        ('Total MRR', 
         ["=B{r}+B{s}".format(r=table_row+3, s=table_row+4)] * 5,
         FINANCE_FORMATS['currency'], True),
        ('Professional ARR', 
         ["=B{r}*12".format(r=table_row+3)] * 5,
         FINANCE_FORMATS['currency'], False),
        ('Enterprise ARR', 
         ["=B{r}*12".format(r=table_row+4)] * 5,
         FINANCE_FORMATS['currency'], False),
        ('Total ARR', 
         ["=B{r}+B{s}".format(r=table_row+7, s=table_row+8)] * 5,
         FINANCE_FORMATS['currency'], True),
        ('ARPU (Blended)', 
         ["=IFERROR(B{r}/('User Forecast'!D67+'User Forecast'!E68),0)".format(r=table_row+5)] * 5,
         FINANCE_FORMATS['currency'], False),
        ('ARPA Professional', 
         [str(PRICING['professional_monthly'])] * 5,
         FINANCE_FORMATS['currency'], False),
        ('ARPA Enterprise', 
         [str(PRICING['enterprise_monthly'])] * 5,
         FINANCE_FORMATS['currency'], False),
    ]
    
    # Fix Enterprise Users references - use correct cells from User Forecast
    revenue_items[1] = ('Enterprise Users (Avg)', 
         ["=INT('User Forecast'!D67*{ent_mix})".format(ent_mix=PRICING['enterprise_mix']/PRICING['professional_mix'])] * 5,
         FINANCE_FORMATS['integer'], False)
    
    for i, (item_name, formulas, fmt, is_total) in enumerate(revenue_items):
        r = table_row + 1 + i
        ws.cell(row=r, column=2, value=item_name)
        ws.cell(row=r, column=2).font = font_body() if not is_total else font_subheader()
        
        for j, formula in enumerate(formulas):
            cell = ws.cell(row=r, column=3+j, value=formula)
            cell.number_format = fmt
            apply_finance_style(cell)
        
        if is_total:
            style_total_row(ws, r, 2, 7)
        else:
            style_data_row(ws, r, 2, 7, i)
    
    # Revenue Recognition Section
    rr_row = table_row + len(revenue_items) + 2
    ws.cell(row=rr_row, column=2, value="Revenue Recognition & Upsell")
    ws.cell(row=rr_row, column=2).font = font_subheader()
    
    rr_headers = ['Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    rr_row += 1
    for col_idx, h in enumerate(rr_headers, start=2):
        ws.cell(row=rr_row, column=col_idx, value=h)
    style_header_row(ws, rr_row, 2, 7)
    
    rr_items = [
        ('Base Subscription Revenue', ["=B{r}*12".format(r=table_row+5)] * 5, FINANCE_FORMATS['currency'], False),
        ('Upsell/Cross-sell (% of Base)', [0.05, 0.08, 0.10, 0.08, 0.06], FINANCE_FORMATS['pct'], False),
        ('Upsell Revenue', ["=B{r}*C{s}".format(r=rr_row+1, s=rr_row+2)] * 5, FINANCE_FORMATS['currency'], False),
        ('Total Recognized Revenue', ["=B{r}+B{s}".format(r=rr_row+1, s=rr_row+3)] * 5, FINANCE_FORMATS['currency'], True),
    ]
    
    for i, (item_name, formulas, fmt, is_total) in enumerate(rr_items):
        r = rr_row + 1 + i
        ws.cell(row=r, column=2, value=item_name)
        ws.cell(row=r, column=2).font = font_body() if not is_total else font_subheader()
        
        for j, formula in enumerate(formulas):
            cell = ws.cell(row=r, column=3+j, value=formula)
            cell.number_format = fmt
            apply_finance_style(cell, is_input=(i == 1))  # Upsell % is input
        
        if is_total:
            style_total_row(ws, r, 2, 7)
        else:
            style_data_row(ws, r, 2, 7, i)
    
    # Column widths
    ws.column_dimensions['B'].width = 28
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 15
    
    return ws


def write_pl_statement(wb):
    """Sheet 4: P&L Statement (5-Year)."""
    ws = wb['P&L Statement']
    last_col = 11
    
    setup_sheet(ws, title="Profit & Loss Statement - 5 Year Projection", last_col=last_col)
    
    # Headers
    row = 4
    headers = ['Line Item', 'Assumption', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    for col_idx, h in enumerate(headers, start=2):
        ws.cell(row=row, column=col_idx, value=h)
    style_header_row(ws, row, 2, 8)
    
    # P&L Structure - Each tuple: (name, formula, assumption, is_section_or_total)
    pl_structure = [
        # Section: REVENUE
        ('REVENUE', None, None, True),
        ('Subscription Revenue', "='Revenue Model'!B38+'Revenue Model'!B39", None, False),
        ('Enterprise Revenue', "='Revenue Model'!B40*0.1", None, False),  # Premium enterprise services
        ('Services/Other Revenue', "=B13*0.02", None, False),
        ('Total Revenue', '=SUM(B11:B13)', None, True),
        
        # Section: COGS
        ('COST OF GOODS SOLD', None, None, True),
        ('Infrastructure (Cloud/Hosting)', '=B14*0.12', '12% of Rev', False),
        ('Customer Support', '=B14*0.05', '5% of Rev', False),
        ('Dev Tools & Licenses', '=B14*0.02', '2% of Rev', False),
        ('Payment Processing', '=B14*0.02', '2% of Rev', False),
        ('Total COGS', '=SUM(B17:B20)', None, True),
        
        # Gross Profit
        ('GROSS PROFIT', '=B14-B21', None, True),
        ('Gross Margin %', '=IFERROR(B22/B14,0)', None, False),
        
        # Section: OPERATING EXPENSES
        ('OPERATING EXPENSES', None, None, True),
        ('Sales & Marketing', '=B14*(0.40-(0.03*(COLUMN()-4)))', '40%→25%', False),
        ('Research & Development', '=B14*(0.50-(0.05*(COLUMN()-4)))', '50%→30%', False),
        ('General & Administrative', '=B14*0.15', '15%', False),
        ('Total Operating Expenses', '=SUM(B27:B29)', None, True),
        
        # EBITDA
        ('EBITDA', '=B22-B30', None, True),
        ('EBITDA Margin %', '=IFERROR(B31/B14,0)', None, False),
        
        # Below Line
        ('Depreciation & Amortization', '=200000+(COLUMN()-4)*100000', None, False),
        ('Interest Expense', '=B33*0.05', '5% of D&A', False),
        ('EBT (Earnings Before Tax)', '=B31-B33-B34', None, True),
        ('Tax Provision (21%)', '=MAX(0,B35*0.21)', '21%', False),
        ('NET INCOME', '=B35-B36', None, True),
        ('Net Margin %', '=IFERROR(B37/B14,0)', None, False),
    ]
    
    current_section_row = None
    item_idx = 0
    
    for i, (item_name, formula, assumption, is_section_or_total) in enumerate(pl_structure):
        r = row + 1 + i
        cell_name = ws.cell(row=r, column=2, value=item_name)
        
        # Style sections and totals differently
        if is_section_or_total:
            if item_name.isupper():  # Section header
                cell_name.font = font_subheader()
                cell_name.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
                for c in range(3, 9):
                    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
                current_section_row = r
            else:  # Total row
                cell_name.font = font_subheader()
                style_total_row(ws, r, 2, 8)
        else:
            cell_name.font = font_body()
            style_data_row(ws, r, 2, 8, item_idx)
            item_idx += 1
        
        # Assumption column
        if assumption:
            cell_assump = ws.cell(row=r, column=3, value=assumption)
            cell_assump.font = font_caption()
        
        # Formula columns (Year 1-5)
        for yr in range(5):
            col = 4 + yr
            cell = ws.cell(row=r, column=col)
            
            if formula:
                # Adjust formula for column (B->C->D->E->F->G for years 1-5)
                col_letter = get_column_letter(col)
                yr_formula = formula.replace('B', col_letter)
                cell.value = yr_formula
            else:
                cell.value = None
            
            # Number format
            if 'Margin' in item_name or '%' in item_name or item_name.startswith('Tax'):
                cell.number_format = FINANCE_FORMATS['pct']
            elif item_name not in [item[0] for item in pl_structure if item[3] and item[0].isupper()]:
                cell.number_format = FINANCE_FORMATS['currency']
            
            apply_finance_style(cell)
    
    # Set column widths
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    for c in range(4, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16
    
    return ws


def write_headcount_plan(wb):
    """Sheet 5: Headcount Plan by Department."""
    ws = wb['Headcount Plan']
    last_col = 8
    
    setup_sheet(ws, title="Headcount Plan - 5 Year Projection", last_col=last_col)
    
    # Salary Assumptions
    row = 4
    ws.cell(row=row, column=2, value="Salary & Benefits Assumptions")
    ws.cell(row=row, column=2).font = font_subheader()
    
    salary_info = [
        ('Engineering Avg Salary', HEADCOUNT['engineering_avg_salary'], FINANCE_FORMATS['currency']),
        ('Sales Avg Salary', HEADCOUNT['sales_avg_salary'], FINANCE_FORMATS['currency']),
        ('Marketing Avg Salary', HEADCOUNT['marketing_avg_salary'], FINANCE_FORMATS['currency']),
        ('Operations Avg Salary', HEADCOUNT['ops_avg_salary'], FINANCE_FORMATS['currency']),
        ('G&A Avg Salary', HEADCOUNT['ga_avg_salary'], FINANCE_FORMATS['currency']),
        ('Benefits Burden', HEADCOUNT['benefits_burden'], FINANCE_FORMATS['pct']),
    ]
    
    row += 1
    for i, (name, val, fmt) in enumerate(salary_info):
        ws.cell(row=row+i, column=2, value=name)
        cell = ws.cell(row=row+i, column=3, value=val)
        apply_finance_style(cell, is_input=True)
        cell.number_format = fmt
    
    # Headcount Table
    table_row = row + len(salary_info) + 2
    headers = ['Department', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    for col_idx, h in enumerate(headers, start=2):
        ws.cell(row=table_row, column=col_idx, value=h)
    style_header_row(ws, table_row, 2, 7)
    
    # Headcount projections (headcounts grow with company scale)
    headcount_data = [
        ('Engineering', [15, 40, 90, 200, 400]),
        ('Sales', [8, 25, 60, 150, 350]),
        ('Marketing', [5, 15, 35, 80, 180]),
        ('Operations', [4, 12, 30, 70, 150]),
        ('General & Admin', [4, 10, 20, 40, 80]),
        ('TOTAL HEADCOUNT', None),  # Will be SUM formula
    ]
    
    for i, (dept, counts) in enumerate(headcount_data):
        r = table_row + 1 + i
        ws.cell(row=r, column=2, value=dept)
        
        if dept == 'TOTAL HEADCOUNT':
            ws.cell(row=r, column=2).font = font_subheader()
            for j in range(5):
                col_letter = get_column_letter(3 + j)
                cell = ws.cell(row=r, column=3+j, value=f"=SUM({col_letter}{table_row+1}:{col_letter}{table_row+5})")
                cell.number_format = FINANCE_FORMATS['integer']
                apply_finance_style(cell)
            style_total_row(ws, r, 2, 7)
        else:
            for j, count in enumerate(counts):
                cell = ws.cell(row=r, column=3+j, value=count)
                cell.number_format = FINANCE_FORMATS['integer']
                apply_finance_style(cell, is_input=True)
            style_data_row(ws, r, 2, 7, i)
    
    # Personnel Costs Section
    cost_row = table_row + len(headcount_data) + 2
    ws.cell(row=cost_row, column=2, value="Personnel Cost Calculation")
    ws.cell(row=cost_row, column=2).font = font_subheader()
    
    cost_headers = ['Cost Category', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    cost_row += 1
    for col_idx, h in enumerate(cost_headers, start=2):
        ws.cell(row=cost_row, column=col_idx, value=h)
    style_header_row(ws, cost_row, 2, 7)
    
    cost_items = [
        ('Engineering Cost', f"=$B${table_row+1}*$B$6*(1+$B$11)", FINANCE_FORMATS['currency']),
        ('Sales Cost', f"=$B${table_row+2}*$B$7*(1+$B$11)", FINANCE_FORMATS['currency']),
        ('Marketing Cost', f"=$B${table_row+3}*$B$8*(1+$B$11)", FINANCE_FORMATS['currency']),
        ('Operations Cost', f"=$B${table_row+4}*$B$9*(1+$B$11)", FINANCE_FORMATS['currency']),
        ('G&A Cost', f"=$B${table_row+5}*$B$10*(1+$B$11)", FINANCE_FORMATS['currency']),
        ('Total Personnel Cost', '=SUM(B{}:B{})', FINANCE_FORMATS['currency'], True),
    ]
    
    for i, item in enumerate(cost_items):
        cr = cost_row + 1 + i
        ws.cell(row=cr, column=2, value=item[0])
        
        is_total = len(item) > 3 and item[3]
        
        for j in range(5):
            col_letter = get_column_letter(3 + j)
            hc_row = table_row + 1 + i if i < 5 else ''
            
            if is_total:
                formula = f"=SUM({col_letter}{cost_row+1}:{col_letter}{cost_row+5})"
            else:
                # Reference headcount * salary * (1+benefits)
                salary_row = 6 + i  # Row where salary is defined
                benefits_row = 11
                formula = f"={col_letter}{table_row+1+i}*$B${salary_row}*(1+$B${benefits_row})"
            
            cell = ws.cell(row=cr, column=3+j, value=formula)
            cell.number_format = item[1]
            apply_finance_style(cell)
        
        if is_total:
            style_total_row(ws, cr, 2, 7)
        else:
            style_data_row(ws, cr, 2, 7, i)
    
    # Hiring Timeline
    hire_row = cost_row + len(cost_items) + 2
    ws.cell(row=hire_row, column=2, value="Quarterly Hiring Plan (Year 1)")
    ws.cell(row=hire_row, column=2).font = font_subheader()
    
    q_headers = ['Department', 'Q1', 'Q2', 'Q3', 'Q4', 'Total']
    hire_row += 1
    for col_idx, h in enumerate(q_headers, start=2):
        ws.cell(row=hire_row, column=col_idx, value=h)
    style_header_row(ws, hire_row, 2, 7)
    
    quarterly_hiring = [
        ('Engineering', [3, 4, 4, 4]),
        ('Sales', [1, 2, 2, 3]),
        ('Marketing', [1, 1, 1, 2]),
        ('Operations', [1, 1, 1, 1]),
        ('G&A', [1, 1, 1, 1]),
    ]
    
    for i, (dept, quarters) in enumerate(quarterly_hiring):
        hr = hire_row + 1 + i
        ws.cell(row=hr, column=2, value=dept)
        for j, q_count in enumerate(quarters):
            cell = ws.cell(row=hr, column=3+j, value=q_count)
            cell.number_format = FINANCE_FORMATS['integer']
            apply_finance_style(cell, is_input=True)
        # Total formula
        cell = ws.cell(row=hr, column=7, value=f"=SUM(C{hr}:F{hr})")
        cell.number_format = FINANCE_FORMATS['integer']
        apply_finance_style(cell)
        style_data_row(ws, hr, 2, 7, i)
    
    # Column widths
    ws.column_dimensions['B'].width = 24
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 12
    
    return ws


def write_cash_flow(wb):
    """Sheet 6: Cash Flow Statement."""
    ws = wb['Cash Flow']
    last_col = 8
    
    setup_sheet(ws, title="Cash Flow Statement - 5 Year Projection", last_col=last_col)
    
    # Headers
    row = 4
    headers = ['Cash Flow Item', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    for col_idx, h in enumerate(headers, start=2):
        ws.cell(row=row, column=col_idx, value=h)
    style_header_row(ws, row, 2, 7)
    
    # Cash Flow structure
    cf_items = [
        # Operating Activities
        ('OPERATING ACTIVITIES', None, True, True),
        ('Net Income', "='P&L Statement'!B37", False, False),
        ('Depreciation & Amortization', "='P&L Statement'!B33", False, False),
        ('Change in Working Capital', "=-(B14*0.05)", False, False),  # 5% of AR
        ('Cash from Operations', '=SUM(B7:B9)', False, True),
        
        # Investing Activities
        ('INVESTING ACTIVITIES', None, True, True),
        ('Capital Expenditures', "=-500000-(COLUMN()-3)*200000", False, False),
        ('PP&E Purchases', "=-200000-(COLUMN()-3)*100000", False, False),
        ('Cash from Investing', '=SUM(B12:B13)', False, True),
        
        # Financing Activities
        ('FINANCING ACTIVITIES', None, True, True),
        ('Equity Raise (Series A/B/C)', [15000000, 30000000, 75000000, 150000000, 0], False, False),
        ('Debt Financing', [0, 0, 25000000, 50000000, 0], False, False),
        ('Debt Principal Repayment', [0, 0, 0, -5000000, -20000000], False, False),
        ('Cash from Financing', '=SUM(B16:B18)', False, True),
        
        # Net Change
        ('NET CHANGE IN CASH', '=B10+B14+B19', False, True),
        
        # Cash Position
        ('Beginning Cash', [10000000, None, None, None, None], False, False),
        ('ENDING CASH POSITION', '=B21+B22', False, True),
        
        # Runway Analysis
        ('RUNWAY ANALYSIS', None, True, True),
        ('Monthly Burn Rate', "=-'P&L Statement'!B31/12", False, False),
        ('Runway (Months)', '=IFERROR(B24/ABS(B25),0)', False, False),
    ]
    
    for i, (item_name, value, is_section, is_total) in enumerate(cf_items):
        r = row + 1 + i
        cell_name = ws.cell(row=r, column=2, value=item_name)
        
        if is_section:
            cell_name.font = font_subheader()
            cell_name.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
            for c in range(3, 8):
                ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
        elif is_total:
            cell_name.font = font_subheader()
            style_total_row(ws, r, 2, 7)
        else:
            style_data_row(ws, r, 2, 7, i)
        
        # Values for each year
        for yr in range(5):
            col = 3 + yr
            cell = ws.cell(row=r, column=col)
            col_letter = get_column_letter(col)
            
            if value is None:
                continue
            elif isinstance(value, list):
                cell.value = value[yr]
            elif isinstance(value, str):
                yr_value = value.replace('B', col_letter)
                cell.value = yr_value
                
                # Special handling for Beginning Cash (references prior year ending)
                if item_name == 'Beginning Cash':
                    if yr == 0:
                        cell.value = value[0]  # Initial cash
                    else:
                        prev_col = get_column_letter(col - 1)
                        cell.value = f"={prev_col}23"  # Prior year ending cash
            
            # Format
            if 'Runway' in item_name:
                cell.number_format = '0.0'
            elif 'Burn' in item_name:
                cell.number_format = FINANCE_FORMATS['currency']
            elif not is_section:
                cell.number_format = FINANCE_FORMATS['currency']
            
            apply_finance_style(cell, is_input=isinstance(value, list) and not is_section)
    
    # Add conditional formatting for runway warning
    runway_row = row + 1 + len(cf_items) - 1
    ws.conditional_formatting.add(
        f'C{runway_row}:G{runway_row}',
        CellIsRule(operator='lessThan', formula=['12'],
                   font=Font(color=ACCENT_NEGATIVE, bold=True),
                   fill=CF_NEGATIVE_FILL)
    )
    
    # Column widths
    ws.column_dimensions['B'].width = 28
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 15
    
    return ws


def write_key_metrics(wb):
    """Sheet 7: Key Metrics & KPIs."""
    ws = wb['Key Metrics & KPIs']
    last_col = 8
    
    setup_sheet(ws, title="Key Metrics & KPIs - Unit Economics", last_col=last_col)
    
    # Headers
    row = 4
    headers = ['KPI / Metric', 'Benchmark', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    for col_idx, h in enumerate(headers, start=2):
        ws.cell(row=row, column=col_idx, value=h)
    style_header_row(ws, row, 2, 8)
    
    # KPI Definitions and Formulas
    kpis = [
        # Unit Economics Section
        ('UNIT ECONOMICS', None, None, True, True),
        
        ('Customer Acquisition Cost (CAC)', 
         "=('P&L Statement'!B27/'User Forecast'!D67)",
         '< $500', False, False),
        ('Lifetime Value (LTV)', 
         "=('Revenue Model'!B9*'P&L Statement'!B22/'P&L Statement'!B14)/(1-'User Forecast'!B14)",
         '> $3,000', False, False),
        ('LTV:CAC Ratio', 
         "=IFERROR(C9/C8,0)",
         '> 3.0x', False, False),
        ('CAC Payback (Months)', 
         "=IFERROR(C8/('Revenue Model'!B9*'P&L Statement'!B22/'P&L Statement'!B14),0)",
         '< 18 mo', False, False),
        
        # Growth Metrics
        ('GROWTH METRICS', None, None, True, True),
        
        ('Monthly Recurring Revenue (MRR)', 
         "='Revenue Model'!B9",
         '-', False, False),
        ('Annual Recurring Revenue (ARR)', 
         "='Revenue Model'!B13",
         '-', False, False),
        ('MRR Growth Rate (MoM)', 
         "=IFERROR((C15-C15)/C15,0)",  # Simplified - would need MoM data
         '> 10%', False, False),
        ('ARR Growth Rate (YoY)', 
         "=IFERROR((D16-C16)/C16,0)",
         '> 50%', False, False),
        
        # Efficiency Metrics
        ('EFFICIENCY METRICS', None, None, True, True),
        
        ('Magic Number', 
         "=IFERROR(('P&L Statement'!D14-'P&L Statement'!C14)/'P&L Statement'!C27,0)",
         '> 0.75', False, False),
        ('Rule of 40 Score', 
         "='Executive Summary'!D13+'P&L Statement'!D26/'P&L Statement'!D14",
         '> 40%', False, False),
        ('Revenue per Employee', 
         "='P&L Statement'!B14/'Headcount Plan'!B19",
         '> $200K', False, False),
        ('Gross Margin', 
         "='P&L Statement'!B22",
         '> 70%', False, False),
        
        # Engagement Metrics
        ('ENGAGEMENT METRICS', None, None, True, True),
        
        ('Free-to-Paid Conversion', 
         "='User Forecast'!D67/'User Forecast'!C67",
         '> 5%', False, False),
        ('Net Dollar Retention', 
         "=1-'User Forecast'!B14",
         '> 100%', False, False),
        ('ARPU (Blended)', 
         "='Revenue Model'!B14",
         '-', False, False),
    ]
    
    for i, (kpi_name, formula, benchmark, is_section, is_total) in enumerate(kpis):
        r = row + 1 + i
        cell_name = ws.cell(row=r, column=2, value=kpi_name)
        
        if is_section:
            cell_name.font = font_subheader()
            cell_name.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
            for c in range(3, 9):
                ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
        elif is_total:
            cell_name.font = font_subheader()
            style_total_row(ws, r, 2, 8)
        else:
            style_data_row(ws, r, 2, 8, i)
        
        # Benchmark column
        if benchmark and not is_section:
            cell_bench = ws.cell(row=r, column=3, value=benchmark)
            cell_bench.font = font_caption()
        
        # Formula columns
        for yr in range(5):
            col = 4 + yr
            cell = ws.cell(row=r, column=col)
            
            if formula and not is_section:
                col_letter = get_column_letter(col)
                yr_formula = formula.replace('C', col_letter).replace('B', col_letter)
                cell.value = yr_formula
                
                # Determine format
                if 'Ratio' in kpi_name or 'Score' in kpi_name:
                    cell.number_format = FINANCE_FORMATS['multiple']
                elif '%' in kpi_name or 'Conversion' in kpi_name or 'Retention' in kpi_name or 'Margin' in kpi_name or 'Growth' in kpi_name:
                    cell.number_format = FINANCE_FORMATS['pct']
                elif 'Months' in kpi_name:
                    cell.number_format = '0.0'
                elif not is_section:
                    cell.number_format = FINANCE_FORMATS['currency']
                
                apply_finance_style(cell)
    
    # Add conditional formatting for LTV:CAC ratio (green if > 3)
    ltvcac_row = row + 10  # LTV:CAC row
    ws.conditional_formatting.add(
        f'D{ltvcac_row}:H{ltvcac_row}',
        CellIsRule(operator='greaterThan', formula=['3'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    
    # Add conditional formatting for Rule of 40 (green if > 40)
    rule40_row = row + 20  # Rule of 40 row
    ws.conditional_formatting.add(
        f'D{rule40_row}:H{rule40_row}',
        CellIsRule(operator='greaterThan', formula=['0.4'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    
    # Column widths
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    for c in range(4, 9):
        ws.column_dimensions[get_column_letter(c)].width = 14
    
    return ws


def write_charts(wb):
    """Sheet 8: Charts Dashboard."""
    ws = wb['Charts']
    last_col = 14
    
    setup_sheet(ws, title="Financial Visualizations", last_col=last_col)
    
    # Create data tables for charts (hidden from main view but referenced)
    
    # Chart 1: Revenue Growth Bar Chart Data
    chart_data_start = 4
    ws.cell(row=chart_data_start, column=2, value="Revenue Data (for charts)")
    ws.cell(row=chart_data_start, column=2).font = font_subheader()
    
    rev_headers = ['Year', 'Subscription', 'Enterprise', 'Total Revenue']
    for col_idx, h in enumerate(rev_headers, start=2):
        ws.cell(row=chart_data_start+1, column=col_idx, value=h)
    style_header_row(ws, chart_data_start+1, 2, 5)
    
    for yr in range(5):
        r = chart_data_start + 2 + yr
        ws.cell(row=r, column=2, value=f"Year {yr+1}")
        ws.cell(row=r, column=3, value=f"='P&L Statement'!{get_column_letter(4+yr)}11")
        ws.cell(row=r, column=4, value=f"='P&L Statement'!{get_column_letter(4+yr)}12")
        ws.cell(row=r, column=5, value=f"='P&L Statement'!{get_column_letter(4+yr)}14")
        for c in range(2, 6):
            ws.cell(row=r, column=c).number_format = FINANCE_FORMATS['currency']
    
    # Create Revenue Bar Chart
    rev_chart = create_bar_chart(style=10, width=16, height=10)
    data = Reference(ws, min_col=3, min_row=chart_data_start+1, max_col=5, max_row=chart_data_start+6)
    cats = Reference(ws, min_col=2, min_row=chart_data_start+2, max_row=chart_data_start+6)
    rev_chart.add_data(data, titles_from_data=True)
    rev_chart.set_categories(cats)
    setup_chart_titles(rev_chart, title="Revenue Growth by Type", y_title="Revenue ($)", x_title="Year")
    apply_chart_colors(rev_chart, [PRIMARY, ACCENT_POSITIVE, ACCENT_WARNING])
    ws.add_chart(rev_chart, "G4")
    
    # Chart 2: User Growth Line Chart Data
    user_data_row = chart_data_start + 9
    ws.cell(row=user_data_row, column=2, value="User Growth Data")
    ws.cell(row=user_data_row, column=2).font = font_subheader()
    
    user_headers = ['Year', 'Free Users', 'Paid Users', 'Total Users']
    for col_idx, h in enumerate(user_headers, start=2):
        ws.cell(row=user_data_row+1, column=col_idx, value=h)
    style_header_row(ws, user_data_row+1, 2, 5)
    
    for yr in range(5):
        r = user_data_row + 2 + yr
        ws.cell(row=r, column=2, value=f"Year {yr+1}")
        ws.cell(row=r, column=3, value=f"='Executive Summary'!{get_column_letter(3+yr)}7")  # Total Users
        ws.cell(row=r, column=4, value=f"='Executive Summary'!{get_column_letter(3+yr)}8")  # Paid Users
        ws.cell(row=r, column=5, value=f"='Executive Summary'!{get_column_letter(3+yr)}7")  # Same for visual
    
    # Create User Growth Line Chart
    user_chart = create_line_chart(style=10, width=16, height=10)
    data = Reference(ws, min_col=3, min_row=user_data_row+1, max_col=4, max_row=user_data_row+6)
    cats = Reference(ws, min_col=2, min_row=user_data_row+2, max_row=user_data_row+6)
    user_chart.add_data(data, titles_from_data=True)
    user_chart.set_categories(cats)
    for series in user_chart.series:
        series.smooth = True
    setup_chart_titles(user_chart, title="User Growth Trajectory", y_title="Users", x_title="Year")
    apply_chart_colors(user_chart, [PRIMARY, ACCENT_POSITIVE])
    ws.add_chart(user_chart, "G18")
    
    # Chart 3: P&L Trend Data
    pl_data_row = user_data_row + 9
    ws.cell(row=pl_data_row, column=2, value="P&L Trend Data")
    ws.cell(row=pl_data_row, column=2).font = font_subheader()
    
    pl_headers = ['Year', 'Revenue', 'Gross Profit', 'EBITDA', 'Net Income']
    for col_idx, h in enumerate(pl_headers, start=2):
        ws.cell(row=pl_data_row+1, column=col_idx, value=h)
    style_header_row(ws, pl_data_row+1, 2, 6)
    
    for yr in range(5):
        r = pl_data_row + 2 + yr
        col = get_column_letter(4 + yr)
        ws.cell(row=r, column=2, value=f"Year {yr+1}")
        ws.cell(row=r, column=3, value=f"='P&L Statement'!{col}14")
        ws.cell(row=r, column=4, value=f"='P&L Statement'!{col}22")
        ws.cell(row=r, column=5, value=f"='P&L Statement'!{col}31")
        ws.cell(row=r, column=6, value=f"='P&L Statement'!{col}37")
        for c in range(3, 7):
            ws.cell(row=r, column=c).number_format = FINANCE_FORMATS['currency']
    
    # Create P&L Trend Combo Chart (Bar + Line)
    pl_chart = create_bar_chart(style=10, width=16, height=10)
    data = Reference(ws, min_col=3, min_row=pl_data_row+1, max_col=6, max_row=pl_data_row+6)
    cats = Reference(ws, min_col=2, min_row=pl_data_row+2, max_row=pl_data_row+6)
    pl_chart.add_data(data, titles_from_data=True)
    pl_chart.set_categories(cats)
    setup_chart_titles(pl_chart, title="P&L Performance Trend", y_title="Amount ($)", x_title="Year")
    apply_chart_colors(pl_chart, [PRIMARY, ACCENT_POSITIVE, ACCENT_WARNING, ACCENT_NEGATIVE])
    ws.add_chart(pl_chart, "G32")
    
    # Chart 4: Unit Economics Data
    unit_data_row = pl_data_row + 9
    ws.cell(row=unit_data_row, column=2, value="Unit Economics Data")
    ws.cell(row=unit_data_row, column=2).font = font_subheader()
    
    unit_headers = ['Year', 'CAC', 'LTV', 'LTV:CAC']
    for col_idx, h in enumerate(unit_headers, start=2):
        ws.cell(row=unit_data_row+1, column=col_idx, value=h)
    style_header_row(ws, unit_data_row+1, 2, 5)
    
    for yr in range(5):
        r = unit_data_row + 2 + yr
        col = get_column_letter(4 + yr)
        ws.cell(row=r, column=2, value=f"Year {yr+1}")
        ws.cell(row=r, column=3, value=f"='Key Metrics & KPIs'!{col}8")
        ws.cell(row=r, column=4, value=f"='Key Metrics & KPIs'!{col}9")
        ws.cell(row=r, column=5, value=f"='Key Metrics & KPIs'!{col}10")
    
    # Create Unit Economics Bar Chart
    unit_chart = create_bar_chart(grouping="clustered", style=10, width=16, height=10)
    data = Reference(ws, min_col=3, min_row=unit_data_row+1, max_col=4, max_row=unit_data_row+6)
    cats = Reference(ws, min_col=2, min_row=unit_data_row+2, max_row=unit_data_row+6)
    unit_chart.add_data(data, titles_from_data=True)
    unit_chart.set_categories(cats)
    setup_chart_titles(unit_chart, title="Unit Economics: CAC vs LTV", y_title="$ Amount", x_title="Year")
    apply_chart_colors(unit_chart, [ACCENT_NEGATIVE, ACCENT_POSITIVE])  # Red for CAC, Green for LTV
    ws.add_chart(unit_chart, "G46")
    
    # Chart 5: Headcount Growth Data
    head_data_row = unit_data_row + 9
    ws.cell(row=head_data_row, column=2, value="Headcount Data")
    ws.cell(row=head_data_row, column=2).font = font_subheader()
    
    head_headers = ['Year', 'Engineering', 'Sales', 'Marketing', 'Ops', 'G&A']
    for col_idx, h in enumerate(head_headers, start=2):
        ws.cell(row=head_data_row+1, column=col_idx, value=h)
    style_header_row(ws, head_data_row+1, 2, 7)
    
    for yr in range(5):
        r = head_data_row + 2 + yr
        col = get_column_letter(3 + yr)
        ws.cell(row=r, column=2, value=f"Year {yr+1}")
        ws.cell(row=r, column=3, value=f"='Headcount Plan'!{col}8")
        ws.cell(row=r, column=4, value=f"='Headcount Plan'!{col}9")
        ws.cell(row=r, column=5, value=f"='Headcount Plan'!{col}10")
        ws.cell(row=r, column=6, value=f"='Headcount Plan'!{col}11")
        ws.cell(row=r, column=7, value=f"='Headcount Plan'!{col}12")
    
    # Create Headcount Stacked Bar Chart
    head_chart = create_bar_chart(grouping="stacked", style=10, width=16, height=10)
    data = Reference(ws, min_col=3, min_row=head_data_row+1, max_col=7, max_row=head_data_row+6)
    cats = Reference(ws, min_col=2, min_row=head_data_row+2, max_row=head_data_row+6)
    head_chart.add_data(data, titles_from_data=True)
    head_chart.set_categories(cats)
    setup_chart_titles(head_chart, title="Headcount Growth by Department", y_title="Employees", x_title="Year")
    apply_chart_colors(head_chart, [PRIMARY, ACCENT_POSITIVE, ACCENT_WARNING, NEUTRAL_600, PRIMARY_LIGHT])
    ws.add_chart(head_chart, "G60")
    
    # Column widths
    ws.column_dimensions['B'].width = 28
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 14
    
    return ws


def main():
    """Main function to generate the complete financial model."""
    print("Creating Endeavor Science (AETH-1) Financial Model...")
    
    # Create workbook
    wb = create_workbook()
    
    # Build each sheet
    print("  [1/8] Building Executive Summary...")
    write_executive_summary(wb)
    
    print("  [2/8] Building User Forecast...")
    write_user_forecast(wb)
    
    print("  [3/8] Building Revenue Model...")
    write_revenue_model(wb)
    
    print("  [4/8] Building P&L Statement...")
    write_pl_statement(wb)
    
    print("  [5/8] Building Headcount Plan...")
    write_headcount_plan(wb)
    
    print("  [6/8] Building Cash Flow Statement...")
    write_cash_flow(wb)
    
    print("  [7/8] Building Key Metrics & KPIs...")
    write_key_metrics(wb)
    
    print("  [8/8] Building Charts...")
    write_charts(wb)
    
    # Set workbook properties
    wb.properties.creator = "Z.ai"
    wb.properties.title = "Endeavor Science (AETH-1) - 5-Year Financial Model"
    wb.properties.description = "Comprehensive 5-year financial projections for Endeavor Science"
    
    # Save workbook
    output_path = "/home/z/my-project/download/AETH-1_Financial_Model.xlsx"
    wb.save(output_path)
    print(f"\n✓ Financial Model saved to: {output_path}")
    
    return output_path


if __name__ == "__main__":
    main()
